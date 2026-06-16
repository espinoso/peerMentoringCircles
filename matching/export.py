"""Write matched circles to an XLSX laid out like the 2025 'Groups Organize by
topics' sheet: circle name in column A on the first member row, members listed
beneath with their full details, a blank row between circles.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .parser import Participant
from .prompt import Group

HEADERS = [
    "Circle",
    "Full Name",
    "E-mail address",
    "Organization",
    "Position",
    "Specific topic",
    "Keywords",
    "Who they want to meet",
    "Objectives",
    "Chosen themes",
    "Multidisciplinary preference",
    "Comments",
]


def _row(p: Participant) -> list[str]:
    return [
        "",  # circle name filled in by caller on the first member row
        p.name,
        p.email,
        p.organization,
        p.position,
        p.topic,
        p.keywords,
        p.who_to_meet,
        "; ".join(p.objectives),
        "; ".join(p.groups),
        p.multidisciplinary,
        p.comments,
    ]


def build_xlsx(groups: list[Group], participants: list[Participant]) -> bytes:
    by_id = {p.id: p for p in participants}

    wb = Workbook()
    ws = wb.active
    ws.title = "Groups by topic"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E7D32")
    circle_font = Font(bold=True, size=12)
    wrap = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for group in groups:
        members = [by_id[i] for i in group.member_ids if i in by_id]
        if not members:
            continue
        for j, p in enumerate(members):
            row = _row(p)
            if j == 0:
                label = group.name
                if group.rationale:
                    label += f"\n({group.rationale})"
                row[0] = label
            ws.append(row)
            r = ws.max_row
            if j == 0:
                ws.cell(row=r, column=1).font = circle_font
            for cell in ws[r]:
                cell.alignment = wrap
                cell.border = border
        ws.append([])  # blank separator row

    widths = [26, 22, 26, 26, 18, 34, 34, 28, 30, 30, 30, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
